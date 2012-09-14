from django.shortcuts import render
from django import forms
from django.http import HttpResponse
from cat.models import MuseumObject
import django_tables2 as tables
from django_tables2 import RequestConfig
from django_tables2.utils import AttributeDict
from django_tables2.utils import A
import refinery
from django.db.models import Q


class ItemTable(tables.Table):
    photo = tables.TemplateColumn('{% load thumbnail %}<img src="{{ record.artefactrepresentation_set.all.0.image|thumbnail_url:\'small_thumb\'  }}" width="105px" height="70px" '
    'alt="">')
    registration_number = tables.LinkColumn(
            'admin:cat_museumobject_change', args=[A('registration_number')])
    category = tables.TemplateColumn('{{ record.categories }}')

    class Meta:
        model = MuseumObject
        sequence = ("photo", "registration_number", "category", "...")


from smart_selects.db_fields import ChainedForeignKey
from smart_selects.form_fields import ChainedModelChoiceField
from smart_selects.widgets import ChainedSelect


class ChainedModelChoiceFilter(refinery.filters.Filter):
    field_class = ChainedModelChoiceField


class ItemFilterSet(refinery.FilterTool):
    def extra(f):
        return {
            'widget': ChainedSelect(f.app_name, f.model_name, f.chain_field, f.model_field, f.show_all, f.auto_choose),
            'app_name': f.app_name,
            'model_name': f.model_name,
            'chain_field': f.chain_field,
            'model_field': f.model_field,
            'show_all': f.show_all,
            'auto_choose': f.auto_choose
        }
    filter_overrides = {
        ChainedForeignKey: {
            'filter_class': ChainedModelChoiceFilter,
            'extra': extra
        }
    }

    def registration_number_filter(values):
        if values:
            values = [int(v) for v in values.split(' ')]
            return Q(registration_number__in=values)
        else:
            return Q()
#    collections = django_filters.ModelChoiceFilter(name='Collection',
#            extra = lambda f: {'queryset':
#             f.rel.to._default_manager.complex_filter(f.rel.limit_choices_to),
#              'to_field_name': f.rel.field_name})
    registration_number = refinery.CharFilter(
            action=registration_number_filter)

    def has_images_filter(value=None):
        if value is not None:
            return Q(artefactrepresentation__isnull=(not value))
        else:
            return Q()

    has_images = refinery.BooleanFilter(
            action=has_images_filter)

    class Meta:
        model = MuseumObject
        fields = ['registration_number', 'functional_category', 'category',
                'artefact_type', 'loan_status', 'access_status',
                'record_status', 'raw_material', 'description',
                'storage_section', 'storage_unit', 'storage_bay',
                'storage_shelf_box_drawer', 'acquisition_date',
                'acquisition_method', 'donor', 'collector',
                'maker', 'has_images', 'global_region', 'country',
                'state_province', 'region_district', 'locality'
                ]


class ColumnForm(forms.Form):
    """For selecting which columns are included in search results"""
    columns = forms.MultipleChoiceField(
            widget=forms.SelectMultiple(attrs={'size': '15'}))
    #widget=forms.CheckboxSelectMultiple)

    def __init__(self, model, *args, **kwargs):
        self.model = model
        super(ColumnForm, self).__init__(*args, **kwargs)
        self.fields['columns'].choices = [(f.name, f.verbose_name.title())
            for f in self.get_all_fields()]

    def get_excluded_names(self):
        """Return list of field names to exclude"""
        desired_fields = self.get_desired_field_names()
        all_fields = self.get_all_fields()

        return [field.name
                for field in all_fields if field.name not in desired_fields]

    def get_desired_fields(self):
        """Return list of fields (not just field name)"""
        desired_fields = self.get_desired_field_names()
        all_fields = self.get_all_fields()

        return [field for field in all_fields if field.name in
                desired_fields]

    def get_desired_field_names(self):
        desired_fields = []
        if self.is_valid():
            desired_fields = self.cleaned_data['columns']
        return desired_fields

    def get_all_fields(self):
        all_fields = [f for f in MuseumObject._meta.fields]
        categories = self.model._meta.many_to_many[0]
        all_fields.insert(5, categories)
        return all_fields


def search_home(request,
        template_name='advanced_search.html'):
    """
    Advanced search/filter form
    """
    is_filtered = bool(request.GET)
    filter = ItemFilterSet(request.GET or None)

    columns = ColumnForm(MuseumObject,
        initial={'columns': ('registration_number', 'artefact_type',
            'category')})
    if request.method == 'GET' and \
            'columns' in request.GET:
        columns = ColumnForm(MuseumObject, request.GET)

    exclude = columns.get_excluded_names()

    # fix field attrs
    for field in filter.form.fields.values():
        if isinstance(field, forms.fields.CharField):
            field.widget.attrs['class'] = 'vTextField'

    title = "Advanced search/filter"
    if 'result' in request.path:
        title = "Search results"

    table = None
    if is_filtered:
        table = ItemTable(filter.qs,
                attrs=AttributeDict({'class': 'paleblue'}),
                exclude=exclude,
                per_page_field='per_page'
            )
        RequestConfig(request, paginate=True).configure(table)

    return render(request, template_name,
            {'filter': filter, 'is_filtered': is_filtered,
                'table': table, 'columns': columns,
                'title': title})


from openpyxl.workbook import Workbook
from openpyxl.style import Color, Font, Alignment, Border, Fill
from openpyxl.cell import get_column_letter


def search_xls(request):
    """Generate xlsx file from search filter"""
    is_filtered = bool(request.GET)
    filter = ItemFilterSet(request.GET or None)

    columns = ColumnForm(MuseumObject, request.GET)

    desired_fields = columns.get_desired_fields()

    wb = create_xlsx_workbook(filter.qs, desired_fields,
            creator=request.user.username, description=str(request.GET))

    response = HttpResponse(
        mimetype='application/'
        'vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'inline; filename=search-results.xlsx'
    wb.save(response)

    return response


def get_col_range(worksheet, fields, fieldname):
    column = get_col_by_fieldname(fields, fieldname)
    return "%s%d:%s%d" % (column, 2, letter, ws.get_highest_row())


def get_col_by_fieldname(fields, fieldname):
    names = [f.name for f in fields]
    try:
        idx = names.index(fieldname) + 1
        column = get_column_letter(idx)
        return column
    except ValueError:
        return None


def create_xlsx_workbook(queryset, fields, creator='Unknown', description=''):
    wb = Workbook()
    wb.properties.description = description
    wb.properties.creator = creator
    ws = wb.get_active_sheet()
    ws.title = "Search Results"

    if queryset.count:
        populate_headers(ws, queryset, fields)
        populate_data(ws, queryset, fields)

        setup_header_cells(ws)
        set_column_widths(ws, fields)

    return wb


def populate_headers(worksheet, queryset, fields):
    worksheet.append([f.verbose_name.title() for f in fields])
    worksheet.append([f.name for f in fields])


def populate_data(worksheet, queryset, fields):
    for row, item in enumerate(queryset):
        row += 1
        for col, field in enumerate(fields):
            cell = worksheet.cell(row=row, column=col)
            #TODO: Fix horrible hack before reusing
            if field.name == 'category':
                cell.value = item.categories()
            else:
                cell.value = unicode(getattr(item, field.name))
            cell.style.alignment.wrap_text = True


def setup_header_cells(worksheet):
    header_cells = worksheet.range('A1:%s1' %
            get_column_letter(worksheet.get_highest_column()))
    for row in header_cells:
        for cell in row:
            cell.style.font.name = 'Arial'
            cell.style.font.bold = True
#    cell.style.fill.fill_type = Fill.FILL_SOLID
#    cell.style.fill.start_color.index = Color.DARKRED
#    cell.style.font.color.index = Color.GREEN

    worksheet.freeze_panes = 'A2'
    worksheet.auto_filter = 'A1:%s1' % (
        get_column_letter(worksheet.get_highest_column()),)


def set_column_widths(worksheet, fields, default_width=15):
    all_columns = [get_column_letter(col + 1) for col in
            range(worksheet.get_highest_column())]
    for column in all_columns:
        worksheet.column_dimensions[column].width = default_width

    field_widths = {
            'acquisition_method': 20,
            'comment': 30,
            'description': 30,
            'donor': 20,
            'collector': 20}
    for name, width in field_widths.iteritems():
        column = get_col_by_fieldname(fields, name)
        if column:
            worksheet.column_dimensions[column].width = width

