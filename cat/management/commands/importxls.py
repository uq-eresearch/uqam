from django.core.management.base import BaseCommand, CommandError
from cat.models import MuseumObject, FunctionalCategory, ArtefactType, CulturalBloc
from cat.models import Person, Place, Maker
from importcat import set_category

from openpyxl.reader.excel import load_workbook

def import_xlsx(filename):
    wb = load_workbook(filename = filename)
    sheet = wb.get_sheet_by_name(name = 'Sheet1')
    sheet.get_highest_column()
    sheet.get_highest_row()
    sheet.cell(row=1, column=1)
    data = sheet.range(sheet.calculate_dimension())
    headers = [v.value for v in data[0]]
    for row in data[1:]:
        vals = [str(v.value).strip() for v in row]
        m = MuseumObject()
        m.id = vals[0]
        m.registration_number = vals[0]
        m.old_registration_number = vals[1]
        m.other_number = vals[2]
        m.functional_category, created = FunctionalCategory.objects.get_or_create(name=vals[3])
        m.artefact_type, created = ArtefactType.objects.get_or_create(name=vals[4])
        m.storage_section = vals[5]
        m.storage_unit = vals[6]
        m.storage_bay = vals[7]
        m.storage_shelf_box_drawer = vals[8]
        m.acquisition_date = vals[9] ####
        m.access_status = vals[10]
        m.description = vals[11]
        m.comment = vals[12]
        region = vals[13]
        state = vals[14]
        m.acquisition_method = vals[15]
        m.loan_status = vals[16]
        country = vals[17]
        place = vals[18]
        m.place, created = Place.objects.get_or_create(name=place,
                                              australian_state=state,
                                              region=region,
                                              country=country)
        m.cultural_bloc, created = CulturalBloc.objects.get_or_create(name=vals[19])
        #TODO: what is photographic_record, is it in Access anywhere
        photographic_record = vals[20]
        m.collector, created = Person.objects.get_or_create(name=vals[21])
        m.when_collector_obtained = vals[22]
        m.how_collector_obtained = vals[23]
        m.donor, created = Person.objects.get_or_create(name=vals[24])
# All empty        when_obtained = vals[25]
# All empty       how_obtained = vals[26]
# All empty       category_illustration = vals[27]
# All empty       artefaction_illustration = vals[28]
        m.raw_material = vals[29]
        m.indigenous_name = vals[30]
        m.assoc_cultural_group = vals[31]
        m.recorded_use = vals[32]
        m.maker_or_artist, created = Maker.objects.get_or_create(name=vals[33])
        def mapint(attr, val):
            try:
                setattr(m, attr, int(val))
            except ValueError:
                pass
        mapint('length', vals[34])
        mapint('width', vals[35])
        mapint('depth', vals[36])
        mapint('circumference', vals[37])
        mapint('weight', vals[38])
        m.save()
        set_category(m, vals[4])




class Command(BaseCommand):
    help = "Import artefact records from xlsx"

    def handle(self, *args, **options):
        if len(args) != 1:
            raise CommandError("need exactly one argument for xlsx file")

        filename, = args

        import_xlsx(filename)


