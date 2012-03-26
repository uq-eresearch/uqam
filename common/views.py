from django.shortcuts import render
from cat.models import MuseumObject, ArtefactType, Category
from django import forms
from django.db.models import Q
from django.http import HttpResponse


# Based on code from:
# http://www.slideshare.net/tpherndon/django-search-presentation

class SearchForm(forms.Form):
    """
    Advanced search form for Museum Objects
    """
    keywords = forms.CharField(
        help_text='Keyword search of item comments and description',
        required=False)
    item_type = forms.ModelChoiceField(queryset=ArtefactType.objects.all(),
            required=False)
    region = forms.CharField(help_text='Region item is from', required=False)
    material = forms.CharField(help_text='Material item is made from',
            required=False)
    people = forms.CharField(help_text='Associated people', required=False)
    category = forms.ModelChoiceField(queryset=Category.objects.all(),
            required=False)
    # maybe dates


def search_home2(request, form_class=SearchForm,
        template_name='common/search_home.html'):
    """
    Advanced search view
    """
    form = None
    if request.method == 'POST':
        #do search
        form = form_class(request.POST)
        if form.is_valid():
            results = search(form.cleaned_data)
            if results:
                return render(template_name, {'form': form, 'items': results})
    else:
        form = form_class()
    return render(request, template_name, {'form': form})


def search(search_data):
    """
    Iterates over all submitted data. For each field, calls
    a method search_'fieldname' on the Searcher.
    """
    q = Q()
    results = None
    searcher = ItemSearch(search_data)

    for key in search_data.iterkeys():
        dispatch = getattr(searcher, 'search_%s' % key)
        q = dispatch(q)

    if q and len(q):
        results = MuseumObject.objects.filter(q).select_related()  # order_by
    else:
        results = []
    return results


class ItemSearch(object):
    def __init__(self, search_data):
        # Include all the form data in self
        self.__dict__.update(search_data)

    def search_keywords(self, q):
        """
        Search should do both title and description, and OR the results
        together, and must iterate over list of keywords
        """
        if self.keywords:
            words = self.keywords.split()
            title_q = Q()
            desc_q = Q()
            for word in words:
                title_q = title_q | Q(title__icontains=word)
                desc_q = desc_q | Q(description__icontains=word)
            keyword_q = title_q | desc_q
            q = q & keyword_q
        return q

    def search_region(self, q):
        return q

    def search_material(self, q):
        return q

    def search_people(self, q):
        return q

    def search_date_added(self, q):
        if self.date_added:
            q = q & Q(date_added__exact=self.date_added)
        return q

    def search_item_type(self, q):
        if self.item_type:
            q = q & Q(artefact_type__icontains=self.artefact_type)
        return q

    def search_category(self, q):
        """
        ManyToMany search
        """
        if self.category:
            if isinstance(self.category, list):
                q = q & Q(category__in=self.category)
            else:
                q = q & Q(category__icontains=self.category)
        return q


import django_filters
import django_tables2 as tables
from django_tables2 import RequestConfig
from django_tables2.utils import AttributeDict
from django_tables2.utils import A


class ItemTable(tables.Table):
    photo = tables.TemplateColumn('<img src="{% firstof '
    'record.artefactrepresentation_set.all.0.image.url_64x64 '
    '"http://placehold.it/48" %}" width="64px" height="64px" '
    'alt="">')
    registration_number = tables.LinkColumn(
            'artefact_view', args=[A('registration_number')])
    class Meta:
        model = MuseumObject
        sequence = ("photo", "registration_number", "...")


class ItemFilterSet(django_filters.FilterSet):
#    collections = django_filters.ModelChoiceFilter(name='Collection',
#            extra = lambda f: {'queryset':
#              f.rel.to._default_manager.complex_filter(f.rel.limit_choices_to),
#              'to_field_name': f.rel.field_name})

    class Meta:
        model = MuseumObject
        fields = ['registration_number', 'functional_category',
                'artefact_type', 'raw_material', 'description', 'category',
                'storage_section', 'storage_unit', 'storage_bay',
                'storage_shelf_box_drawer', 'acquisition_date',
                'acquisition_method', 'cultural_bloc', 'donor', 'collector',
#                'collections__title'
                ]


class ColumnForm(forms.Form):
    """For selecting which columns are included in search results"""
    columns = forms.MultipleChoiceField(
            widget=forms.SelectMultiple(attrs={'size': '15'}))
    #widget=forms.CheckboxSelectMultiple)

    def __init__(self, model, *args, **kwargs):
        self.model = model
        super(ColumnForm, self).__init__(*args, **kwargs)
        self.fields['columns'].choices = [(f.name, f.verbose_name.title())
            for f in self.model._meta.fields]

    def get_excluded_names(self):
        """Return list of field names to exclude"""
        desired_fields = self.get_desired_field_names()
        all_fields = [f for f in self.model._meta.fields]

        return [field.name
                for field in all_fields if field.name not in desired_fields]

    def get_desired_fields(self):
        """Return list of fields (not just field name)"""
        desired_fields = self.get_desired_field_names()
        all_fields = [f for f in MuseumObject._meta.fields]

        return [field for field in all_fields if field.name in
                desired_fields]

    def get_desired_field_names(self):
        desired_fields = []
        if self.is_valid():
            desired_fields = self.cleaned_data['columns']
        return desired_fields


def search_home(request,
        template_name='common/search_home.html'):
    is_filtered = bool(request.GET)
    filter = ItemFilterSet(request.GET or None)

    columns = ColumnForm(MuseumObject)
    if request.method == 'GET' and \
            'columns' in request.GET:
        columns = ColumnForm(MuseumObject, request.GET)

    exclude = columns.get_excluded_names()

    table = None
    if is_filtered:
        table = ItemTable(filter.qs,
                attrs=AttributeDict({'class': 'paleblue'}),
                exclude=exclude
            )
        RequestConfig(request, paginate={"per_page": 30}).configure(table)

    return render(request, template_name,
            {'filter': filter, 'is_filtered': is_filtered,
                'table': table, 'columns': columns})


from openpyxl.workbook import Workbook
from openpyxl.style import Color, Font, Alignment, Border, Fill
from openpyxl.cell import get_column_letter


def search_xls(request):
    """Generate xlsx file from search filter"""
    is_filtered = bool(request.GET)
    filter = ItemFilterSet(request.GET or None)

    columns = ColumnForm(MuseumObject, request.GET)

    desired_fields = columns.get_desired_fields()

    wb = create_xlsx_workbook(filter.qs, desired_fields)

#    response = HttpResponse(mimetype='application/vnd.ms-excel')
    response = HttpResponse(mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'inline; filename=search-results.xlsx'
    wb.save(response)

    return response


def get_col_range(worksheet, fields, fieldname):
    column = get_col_by_fieldname(fields, fieldname)
    return "%s%d:%s%d" % (column, 2, letter, ws.get_highest_row())


def get_col_by_fieldname(fields, fieldname):
    names = [f.name for f in fields]
    try:
        idx = names.index(fieldname) + 1
        column = get_column_letter(idx)
        return column
    except ValueError:
        return None


def create_xlsx_workbook(queryset, fields):
    wb = Workbook()
    ws = wb.get_active_sheet()
    ws.title = "Search Results"

    if queryset.count:
        populate_headers(ws, queryset, fields)
        populate_data(ws, queryset, fields)

        setup_header_cells(ws)
        set_column_widths(ws, fields)

    return wb


def populate_headers(worksheet, queryset, fields):
    worksheet.append([f.verbose_name.title() for f in fields])
    worksheet.append([f.name for f in fields])


def populate_data(worksheet, queryset, fields):
    for row, item in enumerate(queryset):
        row += 1
        for col, field in enumerate(fields):
            cell = worksheet.cell(row=row, column=col)
            cell.value = unicode(getattr(item, field.name))
            cell.style.alignment.wrap_text = True


def setup_header_cells(worksheet):
    header_cells = worksheet.range('A1:%s1' %
            get_column_letter(worksheet.get_highest_column()))
    for row in header_cells:
        for cell in row:
            cell.style.font.name = 'Arial'
            cell.style.font.bold = True
#    cell.style.fill.fill_type = Fill.FILL_SOLID
#    cell.style.fill.start_color.index = Color.DARKRED
#    cell.style.font.color.index = Color.GREEN

    worksheet.freeze_panes = 'A2'
    worksheet.auto_filter = 'A1:%s1' % (
        get_column_letter(worksheet.get_highest_column()),)


def set_column_widths(worksheet, fields, default_width=15):
    all_columns = [get_column_letter(col + 1) for col in
            range(worksheet.get_highest_column())]
    for column in all_columns:
        worksheet.column_dimensions[column].width = default_width

    field_widths = {
            'acquisition_method': 20,
            'comment': 30,
            'description': 30,
            'donor': 20,
            'collector': 20}
    for name, width in field_widths.iteritems():
        column = get_col_by_fieldname(fields, name)
        if column:
            worksheet.column_dimensions[column].width = width

